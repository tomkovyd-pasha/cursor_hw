import pickle
from statistics import mean
import openpyxl
import datetime


with open('task1.txt', 'r') as task_1_file:
    task_1_file_content = task_1_file.readlines()


def list_from_content(is_odd_row: bool) -> list:
    """
    :param is_odd_row: is odd row or not. must be 1 (odd) or 0 (pair)
    :return: list from content
    """
    return eval(f'list(map(lambda y: y.replace("\\n", ""), '
                f'[x for x in task_1_file_content if task_1_file_content.index(x) % 2 {("==", "!=")[is_odd_row]} 0]))')


lst_keys = list_from_content(False)
lst_values = list_from_content(True)

task_1_dict = dict(zip(lst_keys, lst_values))

print(task_1_dict)


def create_str_from_list(lst_to_str: list) -> str:
    """
    create and return string, which return all values from list
    :param lst_to_str: must be list type
    :return: return str
    """
    return ''.join(map(lambda x: ' ' + x, lst_to_str)).strip()


with open('task_1_result.txt', 'w') as task_1_result_file:
    task_1_result_file.write(create_str_from_list(lst_values))


# 2
with open('task2', 'rb') as task_2_file:
    task_2_file_content = pickle.load(task_2_file)
    print(f'Average of task2 file values - {round(mean(task_2_file_content), 2)}')


# 3
class context_manager_openpyxl:
    def __init__(self, filename):
        try:
            self.excel_file = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.excel_file = openpyxl.Workbook()
        finally:
            self.filename = filename

    def __enter__(self):
        return self.excel_file

    def __exit__(self, exc_type, exc_val, exc_trc):
        self.excel_file.save(self.filename)
        self.excel_file.close()
        return True


for i in range(3):
    with context_manager_openpyxl(f'sample{i}.xlsx') as ex_file:
        ex_file_ws = ex_file.create_sheet(0)
        ex_file_ws['A1'] = f'last modified date - {datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'







